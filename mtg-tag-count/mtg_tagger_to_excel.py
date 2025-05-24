#!/usr/bin/env python3
"""
MTG Tag, Type & Deck Export → Single Excel File with Selection, Images & Excel-Based Cumulative Counts

This optimized script parallelizes network calls, caches CSRF tokens, generates a Moxfield import list,
and supports defining both tags and types directly in counts.json.
"""
import argparse
import io
import json
import requests
import shelve
import pandas as pd
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from time import sleep

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# Constants
USER_AGENT = "MTGTagCounter/1.0"
GRAPHQL_URL = "https://tagger.scryfall.com/graphql"
CACHE_PATH = "mtg_cache.db"

# Initialize cache and session
db = shelve.open(CACHE_PATH)
session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})


def get_csrf_token():
    """Fetch CSRF token once and reuse for all GraphQL calls."""
    resp = session.get("https://tagger.scryfall.com")
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    return soup.find("meta", {"name": "csrf-token"})["content"]


# Prepare headers for GraphQL calls
tagger_headers = {"X-CSRF-Token": get_csrf_token(), "Content-Type": "application/json"}

# GraphQL query to fetch taggings
graphql_query = """
query FetchCard($set: String!, $number: String!) {
  card: cardBySet(set: $set, number: $number) {
    taggings(moderatorView: false) {
      tag { name status ancestorTags { name status } }
    }
  }
}
"""


def get_card_info(card_name):
    key = f"info:{card_name.lower()}"
    if key in db:
        return db[key]
    resp = session.get(
        "https://api.scryfall.com/cards/named", params={"exact": card_name}
    )
    resp.raise_for_status()
    info = resp.json()
    db[key] = info
    return info


def get_tagger_tags(set_code, collector_number):
    key = f"tags:{set_code.lower()}:{collector_number}"
    if key in db:
        return set(db[key])
    payload = {
        "query": graphql_query,
        "variables": {"set": set_code, "number": collector_number},
    }
    r = session.post(GRAPHQL_URL, json=payload, headers=tagger_headers)
    r.raise_for_status()
    data = r.json().get("data", {}).get("card", {}) or {}
    tags = set()
    for t in data.get("taggings", []):
        tg = t.get("tag", {})
        if tg.get("status") != "GOOD_STANDING":
            continue
        # Normalize tag name: lowercase, strip, replace spaces with hyphens
        name = tg.get("name", "").lower().strip().replace(" ", "-")
        if name:
            tags.add(name)
        for anc in tg.get("ancestorTags", []):
            if anc.get("status") != "GOOD_STANDING":
                continue
            anc_name = anc.get("name", "").lower().strip().replace(" ", "-")
            if anc_name:
                tags.add(anc_name)
    db[key] = list(tags)
    return tags


def fetch_image_bytes(info):
    if info.get("image_uris"):
        url = info["image_uris"].get("normal")
    elif info.get("card_faces"):
        url = info["card_faces"][0].get("image_uris", {}).get("normal")
    else:
        return None
    r = session.get(url)
    r.raise_for_status()
    return io.BytesIO(r.content)


def fetch_card_data(card, t_tags, t_types):
    info = get_card_info(card)
    raw_tags = get_tagger_tags(info["set"], info["collector_number"])
    types = set(info.get("type_line", "").lower().replace("—", "").split())
    row = {"Image": "", "Card": card}
    for t in t_tags:
        row[t] = t in raw_tags
    for tp in t_types:
        row[tp] = tp in types
    return row, fetch_image_bytes(info), raw_tags


def build_deck_indices(df, rules, exclude_map):
    tally = {k: 0 for k in rules}
    selected = []
    for idx, row in df.iterrows():
        if not any(row.get(k, False) for k in rules):
            continue
        if any(
            rules[k].get("max") is not None
            and row.get(k)
            and tally[k] >= rules[k]["max"]
            for k in rules
        ):
            continue
        primary = [
            k
            for k in rules
            if tally[k] < rules[k]["min"]
            and row.get(k)
            and not any(row.get(ex) for ex in exclude_map.get(k, []))
        ]
        if not primary:
            continue
        selected.append(idx)
        for k in primary:
            tally[k] += 1
        extras = [
            k
            for k in rules
            if k not in primary
            and row.get(k)
            and rules[k].get("max") is not None
            and tally[k] < rules[k]["max"]
        ]
        for k in extras:
            tally[k] += 1
        if all(tally[k] >= rules[k]["min"] for k in rules):
            break
    return selected


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--cards", default="cards.txt")
    parser.add_argument("--tags", default="tags.txt")
    parser.add_argument("--types", default="types.txt")
    parser.add_argument("--counts", default="counts.json")
    parser.add_argument("--output", default="deck.xlsx")
    parser.add_argument("--delay", type=float, default=0.0)
    parser.add_argument("--workers", type=int, default=5)
    args = parser.parse_args()

    # Load inputs
    with open(args.cards) as f:
        cards = [l.strip() for l in f if l.strip()]
    with open(args.tags) as f:
        tracked_tags = [l.strip().lower() for l in f if l.strip()]
    t_tags = tracked_tags
    # Load types from both file and counts.json
    with open(args.types) as f:
        file_types = [l.strip().lower() for l in f if l.strip()]
    counts = json.load(open(args.counts))
    types_from_counts = [k.lower() for k in counts.get("types", {})]
    t_types = sorted(set(file_types + types_from_counts))

    # Exclusion mapping
    exclude_map = {k: list(v) for k, v in counts.get("exclude", {}).items()}

    # Selection rules including both tags and types
    rules = {}
    for k, v in counts.get("tags", {}).items():
        rules[k] = {
            "min": v.get("min", 0) if isinstance(v, dict) else v,
            "max": v.get("max") if isinstance(v, dict) else None,
        }
    for k, v in counts.get("types", {}).items():
        rules[k] = {
            "min": v.get("min", 0) if isinstance(v, dict) else v,
            "max": v.get("max") if isinstance(v, dict) else None,
        }

    # Parallel fetch of card data
    result_map = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(fetch_card_data, c, t_tags, t_types): c for c in cards
        }
        for future in as_completed(futures):
            c = futures[future]
            row, img, raw_tags = future.result()
            result_map[c] = (row, img, raw_tags)
            if args.delay:
                sleep(args.delay)

    rows, imgs, raw_tags_list = [], [], []
    for c in cards:
        r, i, rt = result_map[c]
        rows.append(r)
        imgs.append(i)
        raw_tags_list.append(rt)

    df = pd.DataFrame(rows, index=range(1, len(rows) + 1))
    df.index.name = "Index"

    # Exclusions based on raw tags: clear tutor when tutor-land present
    for idx, rt in enumerate(raw_tags_list, start=1):
        if "tutor" in df.columns and "tutor-land" in rt:
            df.at[idx, "tutor"] = False

    # Build selection and flags
    selected_indices = build_deck_indices(df, rules, exclude_map)
    df["Selected"] = df.index.isin(selected_indices)
    df["NoRelevant"] = ~df[[*t_tags, *t_types]].any(axis=1)

    # Export to Excel and insert images
    output = args.output
    df.to_excel(output, sheet_name="Deck", index_label="Index")
    wb = load_workbook(output)
    ws = wb["Deck"]
    img_col = get_column_letter(df.columns.get_loc("Image") + 2)
    ws.column_dimensions[img_col].width = 20
    for row_idx, img_b in enumerate(imgs, start=2):
        if img_b:
            img = XLImage(img_b)
            img.width, img.height = 160, 224
            img.anchor = f"{img_col}{row_idx}"
            ws.add_image(img)
            ws.row_dimensions[row_idx].height = 180

    # Conditional formatting
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    for col in [*t_tags, *t_types, "Selected", "NoRelevant"]:
        letter = get_column_letter(df.columns.get_loc(col) + 2)
        rng = f"{letter}2:{letter}{len(df) + 1}"
        ws.conditional_formatting.add(rng, CellIsRule("equal", ["TRUE"], green))
        ws.conditional_formatting.add(rng, CellIsRule("equal", ["FALSE"], red))

    # Cumulative COUNTIFS columns for tags
    sel_letter = get_column_letter(df.columns.get_loc("Selected") + 2)
    base_cols = len(df.columns)
    for i, tag in enumerate(t_tags):
        colp = base_cols + 2 + i
        ws.cell(row=1, column=colp, value=f"{tag}_cum")
        orig_letter = get_column_letter(df.columns.get_loc(tag) + 2)
        for r in range(2, len(df) + 2):
            ws.cell(
                row=r,
                column=colp,
                value=f"=COUNTIFS(${sel_letter}$2:${sel_letter}${r},TRUE,${orig_letter}$2:${orig_letter}${r},TRUE)",
            )

    # Create Excel Table
    from openpyxl.worksheet.table import Table, TableStyleInfo

    total_cols = df.shape[1] + 1 + len(t_tags)
    end_col = get_column_letter(total_cols)
    table_ref = f"A1:{end_col}{len(df) + 1}"
    table = Table(displayName="DeckTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(output)

    # Generate Moxfield import
    mox_file = f"{output.rsplit('.', 1)[0]}_moxfield.txt"
    import_lines = []
    for name, count in df[df["Selected"]]["Card"].value_counts().items():
        row = df[df["Card"] == name].iloc[0]
        applied = [f"#{t}" for t in t_tags if row.get(t, False)] + [
            f"#{tp}" for tp in t_types if row.get(tp, False)
        ]
        line = f"{count} {name}" + (f" {' '.join(applied)}" if applied else "")
        import_lines.append(line)
    with open(mox_file, "w") as mf:
        mf.write("\n".join(import_lines))
    print(f"Written Excel to {output} and Mox list to {mox_file}")
    db.close()


if __name__ == "__main__":
    main()
